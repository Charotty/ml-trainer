"""Parse ``Смещение - конечное -12  (2).xlsx`` into canonical Vybor schema."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Mapping, Optional, Sequence

import numpy as np
import pandas as pd

from src.data.excel_displacement_adapter import convert_excel_displacement_df
from src.features.displacement_axis_features import CLINICAL_EXTRA_COLUMNS
from src.features.phase1_schema import TARGET_NAMES, normalize_dataframe

REPO_ROOT = Path(__file__).resolve().parents[2]
_DATA_XLSX = list((REPO_ROOT / "data").glob("*.xlsx"))
DEFAULT_XLSX_PATH = (
    _DATA_XLSX[0] if _DATA_XLSX else REPO_ROOT / "Смещение - конечное -12  (2).xlsx"
)
DEFAULT_OUTPUT_CSV = REPO_ROOT / "data" / "vybor_from_xlsx.csv"

HEADER_ROW = 4
DATA_START_ROW = 5

# 0-based column indices (multi-row header resolved at row HEADER_ROW).
_COL = {
    "row_no": 0,
    "fio": 1,
    "sex": 2,
    "age": 3,
    "body_type": 4,
    "has_previous_surgery": 7,
    "bmi": 8,
    # Col 18 header = "переднезадний (сагиттальный) диаметр" (AP/sagittal) -> depth.
    # Col 19 header = "поперечный диаметр" (transverse) -> width. Previously
    # swapped: body_width_mm held the smaller AP value and body_depth_mm held
    # the larger transverse value, contradicting the width=X/depth=Y
    # convention used everywhere else (see scripts/inference/enhanced_ct_extractor.py).
    "abd_width_l3l4_mm": 19,
    "abd_depth_l3l4_mm": 18,
    "abd_wall_thickness_mm": 11,
    "lumbar_lordosis_deg": 12,
    "s1_plate_tilt_deg": 13,
    "right_kidney_volume_cm3": 20,
    "left_kidney_volume_cm3": 62,
    "right_supine_upper_x": 32,
    "right_supine_upper_y": 33,
    "right_supine_upper_z": 34,
    "right_supine_middle_x": 35,
    "right_supine_middle_y": 36,
    "right_supine_middle_z": 37,
    "right_supine_lower_x": 38,
    "right_supine_lower_y": 39,
    "right_supine_lower_z": 40,
    "right_lateral_upper_x": 47,
    "right_lateral_upper_y": 48,
    "right_lateral_upper_z": 49,
    "right_lateral_middle_x": 50,
    "right_lateral_middle_y": 51,
    "right_lateral_middle_z": 52,
    "right_lateral_lower_x": 53,
    "right_lateral_lower_y": 54,
    "right_lateral_lower_z": 55,
    "left_supine_upper_x": 74,
    "left_supine_upper_y": 75,
    "left_supine_upper_z": 76,
    "left_supine_middle_x": 77,
    "left_supine_middle_y": 78,
    "left_supine_middle_z": 79,
    "left_supine_lower_x": 80,
    "left_supine_lower_y": 81,
    "left_supine_lower_z": 82,
    "left_lateral_upper_x": 89,
    "left_lateral_upper_y": 90,
    "left_lateral_upper_z": 91,
    "left_lateral_middle_x": 92,
    "left_lateral_middle_y": 93,
    "left_lateral_middle_z": 94,
    "left_lateral_lower_x": 95,
    "left_lateral_lower_y": 96,
    "left_lateral_lower_z": 97,
}

_POINTS = ("upper", "middle", "lower")
_SIDES = ("left", "right")


def _parse_numeric(value: object) -> float:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = str(value).replace(",", ".").replace("\u00a0", "").strip()
    if text in {"", "-", "nan", "None"}:
        return np.nan
    try:
        return float(text)
    except ValueError:
        return np.nan


def _cell(row: Sequence[object], col: int) -> object:
    if col >= len(row):
        return None
    return row[col]


def _surname_key(name: object) -> str:
    if name is None:
        return ""
    text = str(name).strip().lower().replace("ё", "е")
    token = re.split(r"[\s.]+", text)[0]
    return re.sub(r"[^a-zа-я0-9]", "", token)


def _delta(lateral: float, supine: float) -> float:
    if not np.isfinite(lateral) or not np.isfinite(supine):
        return np.nan
    return lateral - supine


def _read_xlsx_rows(path: Path) -> list[tuple]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()
    return rows


def parse_xlsx_raw_table(path: Path | str = DEFAULT_XLSX_PATH) -> pd.DataFrame:
    """Return a table compatible with ``train_displacement_dataset.csv`` schema."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"XLSX not found: {path}")

    rows = _read_xlsx_rows(path)
    records: list[dict[str, object]] = []

    for row in rows[DATA_START_ROW:]:
        row_no = _parse_numeric(_cell(row, _COL["row_no"]))
        fio = _cell(row, _COL["fio"])
        if not np.isfinite(row_no) and (fio is None or str(fio).strip() == ""):
            continue

        record: dict[str, object] = {
            "row_no": int(row_no) if np.isfinite(row_no) else len(records) + 1,
            "fio": str(fio).strip() if fio is not None else "",
            "sex": _cell(row, _COL["sex"]),
            "age": _parse_numeric(_cell(row, _COL["age"])),
            "body_type": _cell(row, _COL["body_type"]),
            "has_previous_surgery": _parse_numeric(
                _cell(row, _COL["has_previous_surgery"])
            ),
            "bmi": _parse_numeric(_cell(row, _COL["bmi"])),
            "abd_width_l3l4_mm": _parse_numeric(_cell(row, _COL["abd_width_l3l4_mm"])),
            "abd_depth_l3l4_mm": _parse_numeric(_cell(row, _COL["abd_depth_l3l4_mm"])),
            "abd_wall_thickness_mm": _parse_numeric(_cell(row, _COL["abd_wall_thickness_mm"])),
            "lumbar_lordosis_deg": _parse_numeric(_cell(row, _COL["lumbar_lordosis_deg"])),
            "s1_plate_tilt_deg": _parse_numeric(_cell(row, _COL["s1_plate_tilt_deg"])),
        }

        for side in _SIDES:
            for axis in ("y", "z"):
                sup_u = _parse_numeric(_cell(row, _COL[f"{side}_supine_upper_{axis}"]))
                sup_l = _parse_numeric(_cell(row, _COL[f"{side}_supine_lower_{axis}"]))
                lat_u = _parse_numeric(_cell(row, _COL[f"{side}_lateral_upper_{axis}"]))
                lat_l = _parse_numeric(_cell(row, _COL[f"{side}_lateral_lower_{axis}"]))
                if np.isfinite(sup_u) and np.isfinite(sup_l):
                    record[f"kidney_{side}_{axis}_span_supine_mm"] = abs(sup_u - sup_l)
                if np.isfinite(lat_u) and np.isfinite(lat_l):
                    record[f"kidney_{side}_{axis}_span_lateral_mm"] = abs(lat_u - lat_l)
                sup_span = record.get(f"kidney_{side}_{axis}_span_supine_mm")
                lat_span = record.get(f"kidney_{side}_{axis}_span_lateral_mm")
                if sup_span is not None and lat_span is not None:
                    record[f"kidney_{side}_{axis}_delta_span_mm"] = float(lat_span) - float(sup_span)
            for point in _POINTS:
                for axis in ("x", "y", "z"):
                    sup_key = f"{side}_supine_{point}_{axis}"
                    lat_key = f"{side}_lateral_{point}_{axis}"
                    record[sup_key] = _parse_numeric(_cell(row, _COL[sup_key]))
                    record[lat_key] = _parse_numeric(_cell(row, _COL[lat_key]))
                    record[f"{side}_delta_{point}_{axis}"] = _delta(
                        record[lat_key], record[sup_key]
                    )

        records.append(record)

    return pd.DataFrame(records)


def enrich_with_boku_volumes(
    df: pd.DataFrame,
    boku_path: Path | str,
) -> pd.DataFrame:
    """Fill kidney volumes from na_boku extract when xlsx volume is missing."""
    path = Path(boku_path)
    if not path.exists() or "full_name" not in df.columns and "fio" not in df.columns:
        return df

    boku = pd.read_csv(path, usecols=[
        "case_id",
        "kidney_left_volume_cm3",
        "kidney_right_volume_cm3",
        "kidney_left_length_mm",
        "kidney_right_length_mm",
    ])
    boku["surname_key"] = boku["case_id"].map(_surname_key)
    boku = boku.drop_duplicates("surname_key", keep="first")

    out = df.copy()
    name_col = "full_name" if "full_name" in out.columns else "fio"
    out["surname_key"] = out[name_col].map(_surname_key)
    out = out.merge(
        boku.drop(columns=["case_id"]),
        on="surname_key",
        how="left",
        suffixes=("", "_boku"),
    )
    for col in (
        "kidney_left_volume_cm3",
        "kidney_right_volume_cm3",
        "kidney_left_length_mm",
        "kidney_right_length_mm",
    ):
        boku_col = f"{col}_boku"
        if boku_col in out.columns:
            if col not in out.columns:
                out[col] = out[boku_col]
            else:
                out[col] = out[col].fillna(out[boku_col])
            out = out.drop(columns=[boku_col])
    out = out.drop(columns=["surname_key"], errors="ignore")
    return out


def attach_clinical_extras(converted: pd.DataFrame, raw: pd.DataFrame) -> pd.DataFrame:
    """Merge spine/span columns from raw xlsx table onto canonical rows.

    ``convert_excel_displacement_df`` already sets ``lumbar_lordosis_deg``,
    ``s1_plate_tilt_deg`` and ``abd_wall_thickness_mm`` on ``converted``. A
    plain ``DataFrame.merge`` against the same-named ``extras`` columns from
    ``raw`` used to silently rename BOTH copies to ``<col>_x`` / ``<col>_y``
    (pandas merge-suffix collision), which removed these clinical/anatomical
    features from the canonical schema without raising any error — they were
    present in the CSV under the wrong name and therefore never reached
    ``ANATOMICAL_FEATURES`` / ``DISPLACEMENT_AXIS_FEATURES``. Columns already
    present in ``converted`` are coalesced (NaN-fill only) instead of merged.
    """
    extras = [c for c in CLINICAL_EXTRA_COLUMNS if c in raw.columns]
    if not extras or "fio" not in raw.columns:
        return converted

    raw_key = raw.copy()
    raw_key["_merge_key"] = raw_key["fio"].astype(str).str.strip().str.lower()
    out = converted.copy()
    name_col = "full_name" if "full_name" in out.columns else "fio"
    out["_merge_key"] = out[name_col].astype(str).str.strip().str.lower()
    extra_df = raw_key[["_merge_key"] + extras].drop_duplicates("_merge_key", keep="first")

    already_present = [c for c in extras if c in out.columns]
    new_cols = [c for c in extras if c not in out.columns]

    if new_cols:
        out = out.merge(extra_df[["_merge_key"] + new_cols], on="_merge_key", how="left")

    if already_present:
        rename_map = {c: f"__extra__{c}" for c in already_present}
        coalesce_df = extra_df[["_merge_key"] + already_present].rename(columns=rename_map)
        out = out.merge(coalesce_df, on="_merge_key", how="left")
        for c in already_present:
            shadow_col = rename_map[c]
            out[c] = out[c].fillna(out[shadow_col])
            out = out.drop(columns=[shadow_col])

    return out.drop(columns=["_merge_key"], errors="ignore")


def build_vybor_from_xlsx(
    xlsx_path: Path | str = DEFAULT_XLSX_PATH,
    *,
    boku_path: Optional[Path | str] = None,
    delta_point: str = "middle",
    require_complete_targets: bool = True,
) -> pd.DataFrame:
    """Parse xlsx, convert to Phase-1 schema, optionally enrich from na_boku."""
    raw = parse_xlsx_raw_table(xlsx_path)
    converted = convert_excel_displacement_df(raw, delta_point=delta_point)
    converted = attach_clinical_extras(converted, raw)

    xlsx_rows = _read_xlsx_rows(Path(xlsx_path))
    volumes: list[tuple[float, float]] = []
    for row in xlsx_rows[DATA_START_ROW:]:
        row_no = _parse_numeric(_cell(row, _COL["row_no"]))
        fio = _cell(row, _COL["fio"])
        if not np.isfinite(row_no) and (fio is None or str(fio).strip() == ""):
            continue
        volumes.append(
            (
                _parse_numeric(_cell(row, _COL["left_kidney_volume_cm3"])),
                _parse_numeric(_cell(row, _COL["right_kidney_volume_cm3"])),
            )
        )

    if len(volumes) == len(converted):
        converted["kidney_left_volume_cm3"] = [v[0] for v in volumes]
        converted["kidney_right_volume_cm3"] = [v[1] for v in volumes]

    converted = normalize_dataframe(converted)
    if boku_path:
        converted = enrich_with_boku_volumes(converted, boku_path)

    if require_complete_targets:
        before = len(converted)
        converted = converted.dropna(subset=list(TARGET_NAMES), how="any").copy()
        skipped = before - len(converted)
        if skipped:
            print(
                f"[xlsx] Skipped {skipped} rows with incomplete middle-point deltas "
                f"(kept {len(converted)})"
            )

    converted["source"] = "Vybor"
    converted["source_name"] = "Vybor"
    converted["label_quality"] = "clinical"
    converted["data_origin"] = "Смещение - конечное -12  (2).xlsx"
    return converted


def save_vybor_from_xlsx(
    output_path: Path | str = DEFAULT_OUTPUT_CSV,
    **kwargs: object,
) -> pd.DataFrame:
    df = build_vybor_from_xlsx(**kwargs)  # type: ignore[arg-type]
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_path, index=False)
    print(f"[xlsx] Saved {len(df)} clinical rows -> {output_path}")
    return df
